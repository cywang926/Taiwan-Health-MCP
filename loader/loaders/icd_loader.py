"""
ICD-10-CM 2025 loader.
Parses icd10cm-table-index-2025.zip (NLM format).
Expected files inside ZIP:
  - icd10cm_tabular_2025.xml   (structured tabular format)
  - icd10cm_order_2025.txt     (flat order file — fallback)

ICD-10-PCS 2025 loader.
Parses icd10pcs_tables_2025.zip (CMS format).
Expected file inside ZIP:
  - icd10pcs_codes_2025.txt    (flat: "CODE description", 7-char code + space + text)

Chinese name loader.
Parses the Taiwan MOHW Excel (e.g. 2023年中文版ICD-10-CM_PCS_*.xlsx).
Sheets used: "ICD-10-CM" (col A=code, col D=中文名稱)
             "ICD-10-PCS" (col A=code, col D=中文名稱)
"""

import glob
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from typing import Iterator, Tuple

import asyncpg

# ── Chinese name map from Taiwan MOHW Excel ─────────────────────────────────


def parse_icd_chinese_xlsx(xlsx_path: str) -> tuple[dict[str, str], dict[str, str]]:
    """Parse the Taiwan MOHW bilingual ICD-10 Excel file.

    Reads sheet ``"ICD-10-CM"`` (col 0 = code, col 3 = 中文名稱) and
    ``"ICD-10-PCS"`` (col 0 = 7-char code, col 3 = 中文名稱).

    Args:
        xlsx_path: Path to the MOHW bilingual Excel file.

    Returns:
        A 2-tuple ``(cm_zh, pcs_zh)`` where each element is a dict mapping
        ICD code string to its Chinese name.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to parse the Taiwan ICD bilingual XLSX. "
            "Install project dependencies and rerun the ICD import."
        ) from exc

    print(f"  Parsing Chinese names from {os.path.basename(xlsx_path)} ...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    def _read_sheet(sheet_name: str) -> dict[str, str]:
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not found in xlsx")
            return {}
        ws = wb[sheet_name]
        result: dict[str, str] = {}
        header_skipped = False
        for row in ws.iter_rows(values_only=True):
            if not header_skipped:
                header_skipped = True
                continue
            code = row[0] if len(row) > 0 else None
            name_zh = row[3] if len(row) > 3 else None
            if code and name_zh and isinstance(code, str) and isinstance(name_zh, str):
                result[code.strip()] = name_zh.strip()
        return result

    cm_zh = _read_sheet("ICD-10-CM")
    pcs_zh = _read_sheet("ICD-10-PCS")
    print(f"  Chinese names loaded: {len(cm_zh)} CM, {len(pcs_zh)} PCS")
    return cm_zh, pcs_zh


# ── ICD-10-CM ───────────────────────────────────────────────────────────────


def _parse_xml(data: bytes) -> Iterator[Tuple[str, str, str]]:
    """Yield (code, name_en, name_zh) from NLM tabular XML."""
    root = ET.fromstring(data)

    def walk(node, prefix=""):
        name_node = node.find("name")
        desc_node = node.find("desc")
        if name_node is not None and desc_node is not None:
            code = name_node.text.strip() if name_node.text else ""
            desc = desc_node.text.strip() if desc_node.text else ""
            if code and re.match(r"^[A-Z]\d", code):
                yield code, desc, ""
        for child in node.findall("diag"):
            yield from walk(child)

    for chapter in root.iter("chapter"):
        for section in chapter.findall("section") or [chapter]:
            for diag in section.findall("diag"):
                yield from walk(diag)


def _parse_order_txt(data: bytes) -> Iterator[Tuple[str, str, str]]:
    """
    Fallback: parse icd10cm_order_2025.txt
    Format: seq_no  code  is_header  short_desc  long_desc
    Fields are fixed-width.
    """
    for line in data.decode("utf-8", errors="replace").splitlines():
        if len(line) < 77:
            continue
        code = line[6:13].strip()
        is_header = line[14].strip()
        if is_header == "1":  # header rows don't have billable codes
            continue
        long_desc = line[77:].strip()
        if code and re.match(r"^[A-Z]\d", code):
            yield code, long_desc, ""


async def load_icd10cm(
    pool: asyncpg.Pool,
    zip_path: str,
    name_zh_map: dict[str, str] | None = None,
) -> None:
    """Load ICD-10-CM 2025 codes from the NLM ZIP into ``icd.diagnoses``.

    Prefers the tabular XML; falls back to the flat order TXT if XML is absent.

    Args:
        pool: asyncpg connection pool.
        zip_path: Path to ``icd10cm-table-index-2025.zip``.
        name_zh_map: Optional dict mapping ICD code → Chinese name to include.
    """
    print(f"Parsing {zip_path} ...")
    records: list[tuple] = []
    zh = name_zh_map or {}

    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        xml_files = [
            n for n in names if n.lower().endswith(".xml") and "tabular" in n.lower()
        ]
        txt_files = [
            n for n in names if n.lower().endswith(".txt") and "order" in n.lower()
        ]

        if xml_files:
            print(f"  Using XML: {xml_files[0]}")
            data = zf.read(xml_files[0])
            for code, name_en, _ in _parse_xml(data):
                category = code[:3]
                records.append((code, name_en, zh.get(code, ""), category))
        elif txt_files:
            print(f"  Using TXT: {txt_files[0]}")
            data = zf.read(txt_files[0])
            for code, name_en, _ in _parse_order_txt(data):
                category = code[:3]
                records.append((code, name_en, zh.get(code, ""), category))
        else:
            print(f"  ERROR: No usable file found in {zip_path}. Contents: {names}")
            return

    matched = sum(1 for r in records if r[2])
    print(
        f"  Parsed {len(records)} ICD-10-CM diagnoses ({matched} with Chinese names). Writing to DB ..."
    )

    BATCH = 5000
    async with pool.acquire() as conn:
        await conn.execute("TRUNCATE icd.diagnoses")
        for i in range(0, len(records), BATCH):
            await conn.executemany(
                "INSERT INTO icd.diagnoses (code, name_en, name_zh, category) VALUES ($1,$2,$3,$4)"
                " ON CONFLICT (code) DO UPDATE SET name_en=$2, name_zh=$3, category=$4",
                records[i : i + BATCH],
            )

    print(f"  ICD-10-CM loaded: {len(records)} rows.")
    async with pool.acquire() as conn:
        await conn.execute(
            """INSERT INTO admin.module_load_log (module_key, last_loaded_at, row_count)
               VALUES ('icd', NOW(), $1)
               ON CONFLICT (module_key) DO UPDATE SET last_loaded_at=NOW(), row_count=$1""",
            len(records),
        )


# ── ICD-10-PCS ──────────────────────────────────────────────────────────────


def _parse_pcs_codes_txt(data: bytes) -> Iterator[Tuple[str, str]]:
    """
    Parse CMS icd10pcs_codes_2025.txt
    Format per line: CCCCCCC description
    where CCCCCCC is exactly 7 alphanumeric chars followed by a space.
    """
    for line in data.decode("utf-8", errors="replace").splitlines():
        line = line.rstrip()
        if len(line) < 9:  # 7 code + 1 space + at least 1 char
            continue
        code = line[:7].strip()
        desc = line[8:].strip()
        if code and len(code) == 7 and desc:
            yield code, desc


async def load_icd10pcs(
    pool: asyncpg.Pool,
    zip_path: str,
    name_zh_map: dict[str, str] | None = None,
) -> None:
    print(f"Parsing {zip_path} ...")
    records: list[tuple] = []
    zh = name_zh_map or {}

    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        # Prefer main codes file; exclude addenda (diff/patch files)
        codes_files = [
            n
            for n in names
            if n.lower().endswith(".txt")
            and "addenda" not in n.lower()
            and ("codes" in n.lower() or "pcs" in n.lower())
        ]
        # Fallback: any non-addenda .txt
        if not codes_files:
            codes_files = [
                n
                for n in names
                if n.lower().endswith(".txt") and "addenda" not in n.lower()
            ]

        if not codes_files:
            print(f"  ERROR: No usable PCS file found. Contents: {names}")
            return

        print(f"  Using: {codes_files[0]}")
        data = zf.read(codes_files[0])
        for code, name_en in _parse_pcs_codes_txt(data):
            records.append((code, name_en, zh.get(code, "")))

    matched = sum(1 for r in records if r[2])
    print(
        f"  Parsed {len(records)} ICD-10-PCS procedures ({matched} with Chinese names). Writing to DB ..."
    )

    BATCH = 5000
    async with pool.acquire() as conn:
        await conn.execute("TRUNCATE icd.procedures")
        for i in range(0, len(records), BATCH):
            await conn.executemany(
                "INSERT INTO icd.procedures (code, name_en, name_zh) VALUES ($1,$2,$3)"
                " ON CONFLICT (code) DO UPDATE SET name_en=$2, name_zh=$3",
                records[i : i + BATCH],
            )

    print(f"  ICD-10-PCS loaded: {len(records)} rows.")
